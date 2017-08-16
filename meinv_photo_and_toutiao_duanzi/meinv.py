#_*_ encoding:utf-8 _*_
'''
Created on 2017年8月3日

@author: wangs0622
'''
import urllib
import urllib2
import os
import time
import re
import urlparse
from openpyxl import load_workbook



# 用于下载 http://www.mm131.com 网站上的图片，将下载下来的图片存放在  G:\wechat\meinv 下。
# 因为其网站中一系列的美女图片，只有后面的数字变化，其他不变，所以此程序用于下载某一个系列的图片。
# 例如：http://www.mm131.com/qingchun/3059.html


def changeUrl(url, i, type = 'mm'):
    char_join = '/'
    url_split = url.split(char_join)
    temp = url_split[-1].split('.')
    if type == 'mm':
        temp[0] = str(i)
    else:
        temp[0] = temp[0] + '_' + str(i)
    url_split[-1] = '.'.join(temp)
    return char_join.join(url_split)
        

def download(url, headers={}, proxy=None, num_retries=2, data=None):
    print 'Downloading:', url
    request = urllib2.Request(url, data, headers)
    opener = urllib2.build_opener()
    if proxy:
        proxy_params = {urlparse.urlparse(url).scheme: proxy}
        opener.add_handler(urllib2.ProxyHandler(proxy_params))
    try:
        response = opener.open(request)
        html = response.read()
        code = response.code
    except urllib2.URLError as e:
        print 'Download error:', e.reason
        html = ''
        if hasattr(e, 'code'):
            code = e.code
            if num_retries > 0 and 500 <= code < 600:
                # retry 5XX HTTP errors
                return download(url, headers, proxy, num_retries-1, data)
        else:
            code = None
    return html
    
def main_meinv():
    os.chdir(r'G:\wechat\meinv')
    print "现在的目录为：", os.getcwd()
    url = "http://img1.mm131.com/pic/1036/1.jpg" 
    for i in range(1,16):  
        time.sleep(5)
        url = changeUrl(url, i)
        try:
            urllib.urlretrieve(url,str(i)+'.jpg')
            print "下载第 %d 张图片" %i
        except:
            print url," don't exist "
            break

def getRowNumbers(ws):
    '''
    function:返回sheet表单的行数
    '''
    n = 0
    for row in ws.iter_rows(max_col = 1,min_row = 1):
        if row[0].value is not None:
            n += 1
    return n

def loadexcel(ws):
    exist_duanzi = []

    max_row = getRowNumbers(ws)
    for row in ws.iter_rows(min_row = 1, max_row = max_row, max_col = 30):
        for cell in row:
            if cell.value is not None: exist_duanzi.append(cell.value)
    return exist_duanzi

# 不重复的下载头条段子
#当 like 与 dislike 的比值大于20 的时候才能下载。 

def main_toutiao(url = 'https://www.toutiao.com/api/article/joke/a60745153067/'):
    n = 0
    f = open('result.txt','w')
    wb = load_workbook('existduanzi.xlsx')
    ws = wb.active
    exist_duanzi = loadexcel(ws)
    while n < 30:   # change this number to change the number jokes you want to download.
        header = {}
        header['User_Agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36'
        time.sleep(5)
        html = download(url,header)
        if html is None:
            print "没有下载下来源代码，跳出循环"
            break
           
        regex_like = re.compile("digg_count' : '(\d*)'")
        regex_dislike = re.compile("'bury_count' : '(\d*)'")
        regex_text = re.compile('<div class="content-middle"><p>(.*)</p>')
        regex_nexturl = re.compile('<a ga_event="next_joke" class="right-btn" href="(.*?)"></a>')
        like = regex_like.findall(html)
        dislike = regex_dislike.findall(html)
        text = regex_text.findall(html)
        next_temp = regex_nexturl.findall(html)
        host = urlparse.urlparse(url)
        next_url = host.scheme + '://' + host.netloc + next_temp[0]
        
        number = url.split('/')[-2]
        if number in exist_duanzi:
            print "这个段子下载过了"
        else:
            if int(like[0])/int(dislike[0]) > 20:
                f.write(text[0])
                f.write('\n\n')
                n += 1
                print "下载了第  %d 个 段子" %n
                exist_duanzi.append(number) 
            else:
                print "这个段子赞数不够"
                
        url = next_url
    f.close()
    
    length = len(exist_duanzi)
    if length % 30 > 0:
        max_row = length / 30 +1
    else:
        max_row = length / 30
        
    for row in ws.iter_rows(min_row = 1, max_row = max_row, max_col = 30):
        for cell in row: 
            if len(exist_duanzi) > 0:
                cell.value = exist_duanzi.pop()
            else: break
    wb.save('existduanzi.xlsx')
               
    
            

        
        
if __name__ == '__main__':
    main_toutiao()

    
    
    
